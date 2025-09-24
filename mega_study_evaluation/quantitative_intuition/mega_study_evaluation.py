#!/usr/bin/env python3
"""
Meta-analysis script for Quantitative Intuition
Converted from: quantitative intuition meta analysis.ipynb
"""

import json
import os
import re
import sys
import time
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import f, norm, pearsonr, ttest_rel

# Suppress warnings for cleaner output
warnings.filterwarnings("ignore")

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from common.args_parser import create_base_parser, handle_file_discovery
from common.data_loader import load_standard_data, merge_twin_data, prepare_data_for_analysis
from common.stats_analysis import compute_standard_stats
from common.results_processor import create_results_dict, make_long_format, save_standard_outputs
from common.variable_mapper import create_domain_maps, add_min_max_columns, check_condition_variable, build_min_max_maps
from parse_specification import extract_specification_from_path, parse_specification_name


def process_quantitative_intuition_data(df_human, df_twin, output_path, specification_type):
    """
    Study-specific data processing for quantitative intuition.
    Includes filtering, variable creation, and OpenAI classification.
    """
    # Constants from the original script
    Q_INDIV_COLS = [f"Q20_{i}" for i in range(1, 20) if i != 9]  # drop Q20_9
    I_INDIV_COLS = [f"Q1_{i}" for i in range(1, 20)]
    Q_ORG_COLS = [f"Q22_{i}" for i in range(1, 8)]
    I_ORG_COLS = [f"Q21_{i}" for i in range(1, 10)]  # Q21_10 later excluded
    ATTN_CHECKS_REQUIRED = {"Q20_9": "Somewhat disagree", "Q32": "-3"}  # humans
    ATTN_CHECKS_SKIPPED = ["Q21_10"]
    
    STRAIGHT_LINE_COLS = [f"Q20_{i}" for i in range(13, 20)]
    DURATION_COL = "Duration (in seconds)"
    MIN_DURATION = 150
    LIKERT_MAP = {
        "Strongly disagree": 1,
        "Somewhat disagree": 2,
        "Neither agree nor disagree": 3,
        "Somewhat agree": 4,
        "Strongly agree": 5,
    }

    def clean_tid(df):
        df["_tid"] = pd.to_numeric(df["TWIN_ID"], errors="coerce")
        df = df[df["_tid"].notnull()].copy()
        df["TWIN_ID"] = df["_tid"].astype("Int64")
        return df.drop(columns="_tid")

    def load_qi(df, label: str, apply_filters: bool):
        df = clean_tid(df)
        stats = {
            "dataset": label,
            "initial_n": len(df),
            "removed_attention": 0,
            "removed_straight": 0,
            "removed_duration": 0,
        }

        if DURATION_COL in df.columns:
            df["_duration"] = pd.to_numeric(df[DURATION_COL], errors="coerce")
        else:
            df["_duration"] = np.nan

        likert = list(
            dict.fromkeys(
                Q_INDIV_COLS
                + I_INDIV_COLS
                + Q_ORG_COLS
                + I_ORG_COLS
                + (list(ATTN_CHECKS_REQUIRED) if apply_filters else [])
                + ATTN_CHECKS_SKIPPED
            )
        )
        likert = [c for c in likert if c in df.columns]
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=FutureWarning)
            if likert:
                df[likert] = (
                    df[likert].apply(lambda s: s.astype(str).str.strip()).replace(LIKERT_MAP)
                )

        if apply_filters:
            before = len(df)
            for col, val in ATTN_CHECKS_REQUIRED.items():
                if col in df.columns:
                    expected = LIKERT_MAP.get(val, val)

                    if col == "Q32":
                        col_str = df[col].astype(str).str.strip().str.strip('"')
                        df = df[col_str == str(expected)]
                    else:
                        df = df[df[col].astype(str).str.strip() == str(expected)]

            stats["removed_attention"] = before - len(df)

            before = len(df)
            if all(c in df.columns for c in STRAIGHT_LINE_COLS):
                df = df[df.apply(lambda r: pd.Series(r[STRAIGHT_LINE_COLS]).nunique() > 1, axis=1)]
            stats["removed_straight"] = before - len(df)

            before = len(df)
            if df["_duration"].notna().any():
                df = df[(df["_duration"] >= MIN_DURATION) | df["_duration"].isna()]
            stats["removed_duration"] = before - len(df)

        df["Q_indiv_mean"] = df[[c for c in Q_INDIV_COLS if c in df.columns]].mean(axis=1)
        df["I_indiv_mean"] = df[[c for c in I_INDIV_COLS if c in df.columns]].mean(axis=1)
        df["Q_org_mean"] = df[[c for c in Q_ORG_COLS if c in df.columns]].mean(axis=1)
        df["I_org_mean"] = df[[c for c in I_ORG_COLS if c in df.columns]].mean(axis=1)

        stats["final_n"] = len(df)
        return df, stats

    # Process data with filtering
    df_human, stats_human = load_qi(df_human, "human", apply_filters=True)
    df_twin, stats_twin = load_qi(df_twin, "twin", apply_filters=False)

    # Pair data
    before_pair = len(df_twin)
    df_twin = df_twin[df_twin["TWIN_ID"].isin(df_human["TWIN_ID"])]
    stats_twin.update(
        {
            "dataset": "twin (paired)",
            "initial_n": before_pair,
            "removed_attention": 0,
            "removed_straight": 0,
            "removed_duration": 0,
            "removed_nonpaired": before_pair - len(df_twin),
            "final_n": len(df_twin),
        }
    )

    print("FILTER SUMMARY:")
    print(
        f"  human: start={stats_human['initial_n']} - attn={stats_human['removed_attention']} "
        f"- straight={stats_human['removed_straight']} - duration={stats_human['removed_duration']} "
        f"= final={stats_human['final_n']}"
    )
    print(
        f"  twin : start={stats_twin['initial_n']} - nonpaired_drop={stats_twin['removed_nonpaired']} "
        f"= final={stats_twin['final_n']}"
    )

    # Add close-ended scenario variable
    prefix = "Break the problem into parts"
    for df in (df_human, df_twin):
        df["scenario_close_QI"] = (
            df["Q33"]
            .fillna("")
            .astype(str)
            .str.startswith(prefix)
            .astype(int)
        )

    # Add open-ended scenario classification with OpenAI integration
    FREE_RESP_COL = "Q34"
    CLASSIFICATION_MODEL = "gpt-4o-mini"
    MODEL = CLASSIFICATION_MODEL
    # Cache file should be in the output directory
    # Use specification_type (without date) for cache file names for consistency
    CACHE_FILE = output_path / f"{FREE_RESP_COL}_{specification_type}_open_classification_cache.csv"
    
    prompt_template = (
        "In a research study, a subject was shown quantitative facts about a company, including that "
        "sales are down, call center complaints are up, unresolved complaints are up, and personnel turnover is up. "
        "They were asked to explain what is going on with the company in up to 100 words.\n\n"
        "Classify the SUBJECT ANSWER as either 'summary' (one that just restates or describes the facts) or 'synthesis' "
        "(one that uses a fact to try to explain another, e.g., 'low sales are due to poor customer service', "
        "or provides an alternative explanation altogether, e.g., 'low sales and high complaints are both probably because the product is bad'). "
        'Return JSON with keys classification (summary|synthesis) and confidence (0-100 integer). SUBJECT ANSWER:\n"{ans}"'
    )
    
    # Note: Set OPENAI_API_KEY environment variable or update this line with your key
    OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "YOUR_OPENAI_API_KEY_HERE")
    
    def classify_scenario_open_QI():
        """Classify open-ended scenario responses using OpenAI GPT."""
        # 1) Check cache
        if CACHE_FILE.exists():
            return pd.read_csv(CACHE_FILE)
        
        # 2) Sanity check API key
        if OPENAI_API_KEY.startswith("YOUR_") or not OPENAI_API_KEY:
            print("OpenAI API key not set; skipping scenario_open_QI classification.")
            print("Set OPENAI_API_KEY environment variable or update the script.")
            return pd.DataFrame(columns=["TWIN_ID", "dataset", "scenario_open_QI", "confidence"])
        
        try:
            client = OpenAI(api_key=OPENAI_API_KEY)
            print(f"[DEBUG] Using OpenAI key: {OPENAI_API_KEY[:8]}...{OPENAI_API_KEY[-4:]}")
        except Exception as e:
            print(f"Failed to initialize OpenAI client: {e}")
            return pd.DataFrame(columns=["TWIN_ID", "dataset", "scenario_open_QI", "confidence"])
        
        rows = []
        
        def _process(df, suffix, label):
            col = f"{FREE_RESP_COL}"
            if col not in df.columns:
                return
            subset = df[["TWIN_ID", col]].dropna()
            total = len(subset)
            for i, (tid, ans) in enumerate(subset.itertuples(index=False), 1):
                text = str(ans).strip()
                if not text:
                    continue
                prompt = prompt_template.format(ans=text.replace('"', '\\"'))
                
                backoff = 1
                for attempt in range(4):
                    try:
                        resp = client.chat.completions.create(
                            model=MODEL,
                            messages=[
                                {
                                    "role": "system",
                                    "content": "You are a careful research assistant.",
                                },
                                {"role": "user", "content": prompt},
                            ],
                            temperature=0,
                        )
                        reply = resp.choices[0].message.content
                        m = re.search(r"\{.*\}", reply, re.DOTALL)
                        data = json.loads(m.group(0)) if m else {}
                        cls = data.get("classification", "").lower()
                        # Map summary→0, synthesis→1
                        bin_ = 1 if cls == "synthesis" else 0
                        rows.append((tid, label, bin_, data.get("confidence", None)))
                        break
                    except Exception as e:
                        if attempt == 3:
                            rows.append((tid, label, None, None))
                            print(
                                f"[{label}] ✖ permanent failure at {i}/{total} (TWIN_ID={tid}): {e}"
                            )
                        else:
                            time.sleep(backoff)
                            print(
                                f"[{label}]  attempt {attempt + 1} error at {i}/{total}: {e} — retrying"
                            )
                            time.sleep(backoff)
                            backoff *= 2
                
                if i % 100 == 0:
                    print(f"[{label}] {i}/{total} classified...")
            
            print(f"[{label}] done ({total} items)")
        
        _process(df_human, "human", "human")
        _process(df_twin, "twin", "twin")
        
        out = pd.DataFrame(rows, columns=["TWIN_ID", "dataset", "scenario_open_QI", "confidence"])
        out.to_csv(CACHE_FILE, index=False)
        print(f"Wrote cache → {CACHE_FILE}")
        return out
    
    # Run classification or load from cache
    sc_df = classify_scenario_open_QI()
    
    # Pivot and merge back
    if not sc_df.empty:
        pivot = sc_df.pivot(index="TWIN_ID", columns="dataset", values="scenario_open_QI")
        
        # Map directly into each dataframe
        # For human:
        if "human" in pivot.columns:
            df_human["scenario_open_QI"] = df_human["TWIN_ID"].map(pivot["human"])
        else:
            df_human["scenario_open_QI"] = np.nan
        
        # For twin:
        if "twin" in pivot.columns:
            df_twin["scenario_open_QI"] = df_twin["TWIN_ID"].map(pivot["twin"])
        else:
            df_twin["scenario_open_QI"] = np.nan
    else:
        # Initialize with NaN if no classifications available
        for df in (df_human, df_twin):
            df["scenario_open_QI"] = np.nan

    return df_human, df_twin


def main():
    # Parse arguments - use special configuration for this study
    parser = create_base_parser("Quantitative Intuition")
    args = parser.parse_args()
    
    # Special file handling for quantitative intuition
    study_config = {
        'human_at_study_level': True,
        'human_filename': 'quantitative intuition human data labels anonymized.csv',
        'twin_filename': 'consolidated_llm_labels.csv'
    }
    
    # Handle file discovery
    paths = handle_file_discovery(args, study_config)
    
    if args.verbose:
        print(f"Loading data from:")
        print(f"  Human: {paths['human_data']}")
        print(f"  Twin: {paths['twin_data']}")
        print(f"  Output: {paths['output_dir']}")

    # Study configuration
    study_name = "quantitative intuition"
    
    # Extract specification info
    specification_name = extract_specification_from_path(args.results_dir)
    if specification_name:
        specification_type, ran_date = parse_specification_name(specification_name)
    else:
        specification_name = "unknown"
        specification_type = "unknown" 
        ran_date = None
    
    spec_info = {
        'type': specification_type,
        'name': specification_name,
        'date': ran_date
    }

    # Load and process data
    df_human, df_twin = load_standard_data(paths['human_data'], paths['twin_data'])
    df_human, df_twin = process_quantitative_intuition_data(
        df_human, df_twin, paths['output_path'], specification_type
    )

    # Define study variables  
    condition_vars = [""]  # No conditions
    DV_vars = ["Q_indiv_mean", "I_indiv_mean", "Q_org_mean", "I_org_mean", 
               "scenario_close_QI", "scenario_open_QI"]
    DV_vars_min = [1, 1, 1, 1, 0, 0]
    DV_vars_max = [5, 5, 5, 5, 1, 1]
    
    # Domain classifications (matching original backup for test compatibility)
    DV_vars_social = [0, 0, 0, 0, 0, 0]
    DV_vars_cognitive = [1, 1, 1, 1, 1, 1]
    DV_vars_known = [0, 0, 0, 0, 0, 0]
    DV_vars_pref = [0, 0, 0, 0, 0, 0]
    DV_vars_stim = [0, 0, 0, 0, 0, 0]
    DV_vars_know = [0, 0, 0, 0, 0, 0]
    DV_vars_politics = [0, 0, 0, 0, 0, 0]

    # Merge and prepare data
    df = merge_twin_data(df_human, df_twin)
    df = prepare_data_for_analysis(df, DV_vars)
    
    # Create domain mappings
    domain_maps = create_domain_maps(
        DV_vars, DV_vars_social, DV_vars_cognitive, DV_vars_known, 
        DV_vars_pref, DV_vars_stim, DV_vars_know, DV_vars_politics
    )
    
    # Add min/max columns
    min_map, max_map = build_min_max_maps(DV_vars, DV_vars_min, DV_vars_max)
    df = add_min_max_columns(df, min_map, max_map)
    
    # Check condition variables
    cond_exists, cond, cond_h, cond_t = check_condition_variable(condition_vars)

    # Compute results
    results = []
    for var in DV_vars:
        # Calculate statistics using common function
        stats_dict = compute_standard_stats(
            df, var, cond_exists, cond_h, cond_t, min_map[var], max_map[var]
        )
        
        # Create result dictionary
        result = create_results_dict(
            study_name, var, stats_dict, domain_maps, spec_info
        )
        results.append(result)

    # Convert to DataFrame
    corr_df = pd.DataFrame(results)
    print(corr_df)

    # Create long format data
    df_long = make_long_format(df_human, df_twin, DV_vars, study_name, specification_name)
    print(df_long.head())

    # Save outputs
    save_standard_outputs(corr_df, df_long, paths['output_path'], args.verbose)
    print("done")

    return corr_df, df_long


if __name__ == "__main__":
    main()